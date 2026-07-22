#!/usr/bin/env python
"""Prepare a comprehensive PI/domain-expert review package (v3).

This version is meant to replace the earlier package builders for final review.
It fixes the main package issues:
- explicitly documents meaning-preservation classifier development, feature groups,
  model families tested, and the finalized classifier;
- includes modeling-strategy comparison summaries across all available metrics;
- builds a non-redundant all-in-one Excel workbook for expert review;
- builds source-model crosswalks restricted to DUO/ICO/ODRL/FHIR -> Manual V1 and
  LLM-induced V1, preserving source element IDs/labels/definitions;
- keeps the corrected fixed example HTML from the v2 package when available.

The script does not run any LLMs.
"""
from __future__ import annotations

import argparse
import html
import json
import re
import shutil
from pathlib import Path
from typing import Any

import pandas as pd

try:
    import joblib
except Exception:  # pragma: no cover
    joblib = None  # type: ignore

try:
    import yaml
except Exception:  # pragma: no cover
    yaml = None  # type: ignore

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except Exception as exc:  # pragma: no cover
    raise SystemExit("openpyxl is required. Install with: pip install openpyxl") from exc

SOURCE_MODELS = ["DUO", "ICO", "ODRL", "FHIR"]
SCORE_CANDIDATES = [
    "classifier_preservation_score", "meaning_preserved_score", "meaning_preservation_score",
    "classifier_score", "predicted_probability", "probability", "score",
    "meaning_preserved_pred_proba", "meaning_preserved_pred", "mean_classifier_score",
]


def norm(x: Any) -> str:
    if x is None:
        return ""
    try:
        if pd.isna(x):
            return ""
    except Exception:
        pass
    if isinstance(x, dict):
        return json.dumps(x, ensure_ascii=False)
    if isinstance(x, (list, tuple, set)):
        return "; ".join(norm(v) for v in x if norm(v))
    return " ".join(str(x).split())


def read_csv(path: str | Path | None) -> pd.DataFrame:
    if not path or not Path(path).exists():
        return pd.DataFrame()
    return pd.read_csv(path, low_memory=False).fillna("")


def read_json(path: str | Path | None) -> dict[str, Any]:
    if not path or not Path(path).exists():
        return {}
    try:
        return json.loads(Path(path).read_text())
    except Exception:
        return {}


def read_jsonl(path: str | Path | None) -> list[dict[str, Any]]:
    if not path or not Path(path).exists():
        return []
    rows = []
    with Path(path).open() as f:
        for line in f:
            if line.strip():
                try:
                    rows.append(json.loads(line))
                except Exception:
                    pass
    return rows


def tail(x: Any) -> str:
    s = norm(x)
    return s.split("::", 1)[1] if "::" in s else s


def source_model(x: Any) -> str:
    s = norm(x)
    u = s.upper()
    if "FHIR" in u or s.startswith("Consent."):
        return "FHIR"
    if "ODRL" in u:
        return "ODRL"
    if "DUO" in u:
        return "DUO"
    if "ICO" in u:
        return "ICO"
    if "::" in s:
        return source_model(s.split("::", 1)[0])
    return s


def pick(row: pd.Series | dict[str, Any], cols: list[str]) -> str:
    for c in cols:
        if isinstance(row, pd.Series):
            if c in row.index and norm(row.get(c)):
                return norm(row.get(c))
        elif c in row and norm(row.get(c)):
            return norm(row.get(c))
    return ""


def list_text(x: Any) -> str:
    if isinstance(x, str):
        s = x.strip()
        if not s:
            return ""
        try:
            obj = json.loads(s)
            return list_text(obj)
        except Exception:
            return s
    if isinstance(x, list):
        return "; ".join(norm(v) for v in x if norm(v))
    if isinstance(x, dict):
        return "; ".join(f"{k}: {norm(v)}" for k, v in x.items())
    return norm(x)


def read_schema(path: str | Path | None) -> dict[str, Any]:
    if not path or not Path(path).exists():
        return {}
    p = Path(path)
    txt = p.read_text()
    if p.suffix.lower() == ".json":
        return json.loads(txt)
    if yaml is None:
        raise RuntimeError("PyYAML is required to read YAML schemas.")
    return yaml.safe_load(txt)


def schema_fields(schema: dict[str, Any]) -> list[dict[str, Any]]:
    raw = schema.get("fields") or schema.get("schema_fields") or schema.get("functional_fields") or []
    out = []
    if isinstance(raw, dict):
        for k, v in raw.items():
            d = dict(v) if isinstance(v, dict) else {"definition": v}
            d.setdefault("field_id", k)
            out.append(d)
    elif isinstance(raw, list):
        for i, v in enumerate(raw):
            d = dict(v) if isinstance(v, dict) else {"field_id": str(v)}
            d.setdefault("field_id", d.get("id") or d.get("name") or f"field_{i:02d}")
            out.append(d)
    return out


def source_dictionary(inventory_csv: str | Path, extra_frames: list[pd.DataFrame]) -> pd.DataFrame:
    rows = []
    inv = read_csv(inventory_csv)
    if not inv.empty:
        for _, r in inv.iterrows():
            sid = pick(r, ["union_element_id", "source_element_id", "element_id", "id"])
            sm = source_model(pick(r, ["information_model", "source_model", "model"]) or sid)
            if sm not in SOURCE_MODELS:
                continue
            rows.append({
                "source_model": sm,
                "source_element_id": sid,
                "source_element_tail": tail(sid),
                "source_element_label": pick(r, ["source_element_label", "label", "name", "source_label"]),
                "source_element_definition": pick(r, ["source_element_definition", "definition", "description"]),
                "source_element_uri_or_code": pick(r, ["uri", "url", "code", "concept_id"]),
                "dictionary_source": "source_element_inventory",
            })
    for df in extra_frames:
        if df is None or df.empty:
            continue
        for _, r in df.iterrows():
            sid = pick(r, ["source_element_id", "source_element", "union_element_id", "element_id"])
            sm = source_model(pick(r, ["source_model", "information_model", "model"]) or sid)
            if sm not in SOURCE_MODELS or not sid:
                continue
            rows.append({
                "source_model": sm,
                "source_element_id": sid,
                "source_element_tail": tail(sid),
                "source_element_label": pick(r, ["source_element_label", "source_label", "label", "source_element"]),
                "source_element_definition": pick(r, ["source_element_definition", "definition", "description"]),
                "source_element_uri_or_code": "",
                "dictionary_source": "crosswalk_or_evidence",
            })
    out = pd.DataFrame(rows).drop_duplicates()
    if out.empty:
        return out
    # Keep most informative definition/label per source element.
    out["_info"] = out["source_element_label"].str.len() + out["source_element_definition"].str.len()
    out = out.sort_values("_info", ascending=False).drop_duplicates(["source_model", "source_element_id"]).drop(columns="_info")
    return out.sort_values(["source_model", "source_element_id"])


def make_lookup(src: pd.DataFrame) -> dict[str, dict[str, str]]:
    look: dict[str, dict[str, str]] = {}
    if src.empty:
        return look
    for _, r in src.iterrows():
        d = {c: norm(r.get(c)) for c in src.columns}
        for key in [d.get("source_element_id"), d.get("source_element_tail"), d.get("source_element_label")]:
            if norm(key):
                look.setdefault(norm(key).lower(), d)
    return look


def fill_from_lookup(row: dict[str, Any], look: dict[str, dict[str, str]]) -> dict[str, Any]:
    sid = norm(row.get("source_element_id"))
    match = look.get(sid.lower()) or look.get(tail(sid).lower()) or look.get(norm(row.get("source_element_label")).lower()) or {}
    for c in ["source_model", "source_element_tail", "source_element_label", "source_element_definition"]:
        if not norm(row.get(c)) and match.get(c):
            row[c] = match[c]
    return row


def normalize_manual_crosswalk(manual_crosswalk_csv: str | Path, look: dict[str, dict[str, str]]) -> pd.DataFrame:
    df = read_csv(manual_crosswalk_csv)
    rows = []
    for _, r in df.iterrows():
        sid = pick(r, ["source_element_id", "source_element", "union_element_id", "element_id"])
        sm = source_model(pick(r, ["source_model", "information_model", "model"]) or sid)
        if sm not in SOURCE_MODELS or not sid:
            continue
        d = {
            "source_model": sm,
            "source_element_id": sid,
            "source_element_tail": tail(sid),
            "source_element_label": pick(r, ["source_element_label", "source_label", "label", "source_element"]),
            "source_element_definition": pick(r, ["source_element_definition", "definition", "description"]),
            "manual_v1_field": pick(r, ["manual_v1_field", "v1_field", "functional_field", "proposed_v1_field", "target_field"]),
            "secondary_manual_v1_fields": pick(r, ["secondary_v1_fields_json", "secondary_fields"]),
            "mapping_type": pick(r, ["mapping_type", "manual_mapping_type", "relationship", "mapping_relation"]),
            "rationale_or_context_rule": pick(r, ["manual_rationale", "rationale", "context_rule", "rule", "notes"]),
            "expert_review_status": "",
            "expert_notes": "",
        }
        rows.append(fill_from_lookup(d, look))
    return pd.DataFrame(rows).drop_duplicates()


def load_cards(cards_root: str | Path) -> dict[str, dict[str, Any]]:
    cards: dict[str, dict[str, Any]] = {}
    root = Path(cards_root)
    if not root.exists():
        return cards
    for p in root.glob("fold_*/schema_induction_evidence_cards.jsonl"):
        fold = p.parent.name
        for card in read_jsonl(p):
            ids = [pick(card, ["card_id", "evidence_card_id", "candidate_field_id", "stability_group_id", "id"])]
            ids += [norm(card.get(k)) for k in ["candidate_field_id", "stability_group_id"] if norm(card.get(k))]
            for cid in ids:
                if cid:
                    cards[f"{fold}::{cid}"] = card
                    cards.setdefault(cid, card)
    return cards


def field_evidence_ids(field: dict[str, Any]) -> list[str]:
    ids = []
    for k in ["evidence_card_ids", "assigned_evidence_cards", "supporting_evidence_cards", "evidence_cards", "source_cards", "cluster_ids", "stability_group_ids"]:
        v = field.get(k)
        if isinstance(v, list):
            ids += [norm(x) for x in v if norm(x)]
        elif norm(v):
            ids += [x.strip() for x in re.split(r"[;|,]", norm(v)) if x.strip()]
    text = " ".join(norm(field.get(k)) for k in ["rationale", "evidence", "notes", "definition"])
    ids += re.findall(r"(?:C\d{3,}|SG[_-]?\d+|field[_-]?\d+|cluster[_-]?\d+)", text, flags=re.I)
    return list(dict.fromkeys(ids))


def source_rows_from_card(card: dict[str, Any], look: dict[str, dict[str, str]]) -> list[dict[str, Any]]:
    values = []
    for k in ["top_source_elements", "source_elements", "source_model_elements", "member_elements", "crosswalk_mappings"]:
        v = card.get(k)
        if isinstance(v, list):
            values += v
    rows = []
    for item in values:
        if isinstance(item, dict):
            sid = pick(item, ["source_element_id", "source_element", "union_element_id", "element_id", "id", "label"])
            sm = source_model(pick(item, ["source_model", "information_model", "model"]) or sid)
            lab = pick(item, ["source_element_label", "source_label", "label", "name"])
            definition = pick(item, ["source_element_definition", "definition", "description"])
        else:
            sid = norm(item)
            sm = source_model(sid)
            lab = tail(sid)
            definition = ""
        if sm not in SOURCE_MODELS or not sid:
            continue
        rows.append(fill_from_lookup({
            "source_model": sm,
            "source_element_id": sid,
            "source_element_tail": tail(sid),
            "source_element_label": lab,
            "source_element_definition": definition,
        }, look))
    # If card only has source_models but no source elements, do not create vague rows.
    out = []
    seen = set()
    for r in rows:
        key = (r.get("source_model"), r.get("source_element_id"))
        if key not in seen:
            seen.add(key)
            out.append(r)
    return out


def fallback_cards_for_field(field: dict[str, Any], fold: str, cards: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    # Conservative text overlap fallback so we avoid empty LLM crosswalk rows.
    text = " ".join(norm(field.get(k)) for k in ["field_id", "name", "definition", "description", "examples"]).lower()
    scored = []
    for key, card in cards.items():
        if "::" in key and not key.startswith(fold + "::"):
            continue
        vocab = " ".join(list_text(card.get(k)) for k in ["suggested_terms", "top_spans", "source_models"])
        toks = {t for t in re.findall(r"[a-z0-9_]+", vocab.lower()) if len(t) > 3}
        score = sum(1 for t in toks if t in text)
        if score >= 2:
            scored.append((score, card))
    scored.sort(key=lambda x: x[0], reverse=True)
    return [c for _, c in scored[:3]]


def build_llm_crosswalk(schema_root: str | Path, cards_root: str | Path, look: dict[str, dict[str, str]], manual_xw: pd.DataFrame) -> pd.DataFrame:
    cards = load_cards(cards_root)
    manual_by_source = {}
    if not manual_xw.empty:
        for _, r in manual_xw.iterrows():
            manual_by_source.setdefault(norm(r.get("source_element_id")).lower(), set()).add(norm(r.get("manual_v1_field")))
    rows = []
    root = Path(schema_root)
    if not root.exists():
        return pd.DataFrame()
    for p in sorted(root.glob("fold_*/llm_induced_functional_v1_candidate.*")):
        if p.suffix.lower() not in {".yaml", ".yml", ".json"}:
            continue
        fold = p.parent.name
        schema = read_schema(p)
        for f in schema_fields(schema):
            fid = pick(f, ["field_id", "id", "name"])
            fname = pick(f, ["name", "label", "field_id", "id"]) or fid
            definition = pick(f, ["definition", "description"])
            eids = field_evidence_ids(f)
            used_cards = []
            for eid in eids:
                for key in [f"{fold}::{eid}", eid]:
                    if key in cards:
                        used_cards.append(cards[key])
            if not used_cards:
                used_cards = fallback_cards_for_field(f, fold, cards)
            source_rows = []
            for c in used_cards:
                source_rows += source_rows_from_card(c, look)
            for sr in source_rows:
                sid = norm(sr.get("source_element_id"))
                rows.append({
                    "fold": fold,
                    "source_model": norm(sr.get("source_model")),
                    "source_element_id": sid,
                    "source_element_tail": tail(sid),
                    "source_element_label": norm(sr.get("source_element_label")),
                    "source_element_definition": norm(sr.get("source_element_definition")),
                    "llm_induced_v1_field": fid,
                    "llm_induced_v1_field_name": fname,
                    "llm_induced_v1_definition": definition,
                    "manual_v1_fields_linked_by_source_element": "; ".join(sorted(manual_by_source.get(sid.lower(), set()))),
                    "mapping_basis": "evidence_card_source_elements" if eids else "field_text_to_evidence_card_fallback",
                    "evidence_card_ids": "; ".join(eids),
                    "expert_review_status": "",
                    "expert_notes": "",
                })
    out = pd.DataFrame(rows).drop_duplicates()
    if not out.empty:
        out = out[out["source_model"].isin(SOURCE_MODELS) & out["source_element_id"].astype(str).str.len().gt(0)]
    return out


def manual_dictionary(schema_yaml: str | Path) -> pd.DataFrame:
    schema = read_schema(schema_yaml)
    rows = []
    sd = schema.get("sentence_decision") or {}
    if sd:
        rows.append({"schema": "Manual Functional V1", "field_id": "sentence_decision", "field_name": "sentence_decision", "tier": "core", "definition": pick(sd, ["definition", "description"]) or "Provision-level permit/deny/obligation/mixed/unclear decision.", "include_when": "Provision-level polarity is expressed/inferable.", "exclude_when": "Local negation that does not determine overall provision decision.", "examples": list_text(sd.get("examples")), "expert_review_status": "", "expert_notes": ""})
    for f in schema_fields(schema):
        fid = pick(f, ["field_id", "id", "name"])
        rows.append({"schema": "Manual Functional V1", "field_id": fid, "field_name": pick(f, ["name", "label", "field_id", "id"]) or fid, "tier": pick(f, ["tier", "status", "field_type"]), "definition": pick(f, ["definition", "description"]), "include_when": list_text(f.get("include_when") or f.get("inclusion_criteria") or f.get("inclusion")), "exclude_when": list_text(f.get("exclude_when") or f.get("exclusion_criteria") or f.get("exclusion")), "examples": list_text(f.get("examples") or f.get("example_spans")), "expert_review_status": "", "expert_notes": ""})
    return pd.DataFrame(rows).drop_duplicates()


def llm_dictionary(schema_root: str | Path, consensus_fields_csv: str | Path | None) -> pd.DataFrame:
    rows = []
    root = Path(schema_root)
    if root.exists():
        for p in sorted(root.glob("fold_*/llm_induced_functional_v1_candidate.*")):
            if p.suffix.lower() not in {".yaml", ".yml", ".json"}:
                continue
            fold = p.parent.name
            schema = read_schema(p)
            for f in schema_fields(schema):
                fid = pick(f, ["field_id", "id", "name"])
                rows.append({"schema_level": "fold_specific", "fold": fold, "field_id": fid, "field_name": pick(f, ["name", "label", "field_id", "id"]) or fid, "tier": pick(f, ["tier", "status", "field_type"]), "definition": pick(f, ["definition", "description"]), "include_when": list_text(f.get("include_when") or f.get("inclusion_criteria") or f.get("inclusion")), "exclude_when": list_text(f.get("exclude_when") or f.get("exclusion_criteria") or f.get("exclusion")), "examples": list_text(f.get("examples") or f.get("example_spans")), "evidence_card_ids": "; ".join(field_evidence_ids(f)), "expert_review_status": "", "expert_notes": ""})
    cons = read_csv(consensus_fields_csv)
    if not cons.empty:
        for _, r in cons.iterrows():
            rows.append({"schema_level": "post_cv_consensus", "fold": "consensus", "field_id": pick(r, ["consensus_field", "consensus_field_id", "field_id", "field", "name"]), "field_name": pick(r, ["consensus_field", "consensus_name", "name", "field"]), "tier": pick(r, ["tier", "status", "selection_tier"]), "definition": pick(r, ["definition", "description"]), "include_when": pick(r, ["include_when", "inclusion_criteria"]), "exclude_when": pick(r, ["exclude_when", "exclusion_criteria"]), "examples": pick(r, ["examples", "top_spans", "example_spans"]), "evidence_card_ids": pick(r, ["evidence_card_ids", "supporting_fields", "fold_fields"]), "expert_review_status": "", "expert_notes": ""})
    return pd.DataFrame(rows).drop_duplicates()


def combined_crosswalk(manual_xw: pd.DataFrame, llm_xw: pd.DataFrame) -> pd.DataFrame:
    keys = ["source_model", "source_element_id", "source_element_tail", "source_element_label", "source_element_definition"]
    if manual_xw.empty and llm_xw.empty:
        return pd.DataFrame()
    parts = []
    if not manual_xw.empty:
        parts.append(manual_xw.groupby(keys, dropna=False).agg(manual_v1_fields=("manual_v1_field", lambda x: "; ".join(sorted(set(norm(v) for v in x if norm(v))))), manual_mapping_types=("mapping_type", lambda x: "; ".join(sorted(set(norm(v) for v in x if norm(v)))))).reset_index())
    if not llm_xw.empty:
        parts.append(llm_xw.groupby(keys, dropna=False).agg(llm_induced_v1_fields=("llm_induced_v1_field", lambda x: "; ".join(sorted(set(norm(v) for v in x if norm(v))))), llm_induced_field_names=("llm_induced_v1_field_name", lambda x: "; ".join(sorted(set(norm(v) for v in x if norm(v))))), llm_folds=("fold", lambda x: "; ".join(sorted(set(norm(v) for v in x if norm(v))))), llm_mapping_basis=("mapping_basis", lambda x: "; ".join(sorted(set(norm(v) for v in x if norm(v)))))).reset_index())
    out = parts[0] if len(parts) == 1 else parts[0].merge(parts[1], on=keys, how="outer")
    out["manual_expert_decision"] = ""
    out["llm_expert_decision"] = ""
    out["expert_notes"] = ""
    return out.sort_values(["source_model", "source_element_id"])


def score_column(df: pd.DataFrame) -> str | None:
    return next((c for c in SCORE_CANDIDATES if c in df.columns), None)


def classifier_method_sheet() -> pd.DataFrame:
    rows = [
        ("Goal", "Binary classifier to estimate whether backward reconstruction preserves original sentence meaning."),
        ("Training data", "Human-labeled round-trip outputs from individual information-model experiments."),
        ("Evaluation splits", "Random stratified split, leave-sentence/group folds, leave-one-LLM-out, and leave-one-information-model-out."),
        ("Lexical/content features", "original/reconstructed token length, length ratio, absolute length difference, token Jaccard, TF-IDF cosine."),
        ("Mapping-burden features", "annotation count, unique element count, mapping length, bracket/parenthesis counts."),
        ("Consent cue features", "permission, obligation, prohibition, negation, condition, exception, restriction, withdrawal, action, resource, actor, purpose cue counts/Jaccard/missing/added/presence preservation."),
        ("Modal features", "modal_orig, modal_recon, modal_category_changed for permission/obligation/prohibition/none."),
        ("Semantic features", "embedding cosine/distance and optional NLI entailment/contradiction/neutrality features when configured."),
        ("Model families tested", "majority baseline, bag-of-words TF-IDF logistic regression, structured logistic regression, structured random forest, optional structured XGBoost, optional embedding-only logistic regression."),
        ("Final classifier", "RandomForestClassifier with 500 trees, balanced class weights, min_samples_leaf=2, trained on all labeled rows with selected feature set for scoring new round trips."),
        ("Use in this package", "Classifier score is reported with diagnostic metrics; it is a scalable proxy and should not replace expert review."),
    ]
    return pd.DataFrame(rows, columns=["topic", "details"])


def classifier_artifacts(classifier_dir: str | Path, experiments_dir: str | Path, diagnostics_dir: Path, plots_dir: Path) -> dict[str, pd.DataFrame]:
    out: dict[str, pd.DataFrame] = {}
    cdir = Path(classifier_dir)
    edir = Path(experiments_dir) if experiments_dir else Path("")
    summary = read_json(cdir / "final_classifier_training_summary.json")
    rows = []
    if summary:
        for k, v in summary.items():
            rows.append({"section": "final_training_summary", "metric": k, "value": norm(v)})
    bundle_path = cdir / "final_meaning_preservation_classifier.joblib"
    if bundle_path.exists() and joblib is not None:
        try:
            b = joblib.load(bundle_path)
            rows += [
                {"section": "final_bundle", "metric": "feature_set", "value": norm(b.get("feature_set")) if isinstance(b, dict) else ""},
                {"section": "final_bundle", "metric": "embedding_model", "value": norm(b.get("embedding_model")) if isinstance(b, dict) else ""},
                {"section": "final_bundle", "metric": "nli_model", "value": norm(b.get("nli_model")) if isinstance(b, dict) else ""},
                {"section": "final_bundle", "metric": "n_features", "value": str(len(b.get("feature_columns", []))) if isinstance(b, dict) else ""},
                {"section": "final_bundle", "metric": "model_class", "value": type(b.get("model")).__name__ if isinstance(b, dict) else ""},
            ]
            cue = b.get("cue_dictionary", {}) if isinstance(b, dict) else {}
            for group, cues in (cue.get("cue_groups") or {}).items():
                rows.append({"section": "cue_dictionary", "metric": group, "value": "; ".join(cues)})
            for col in b.get("feature_columns", []) if isinstance(b, dict) else []:
                rows.append({"section": "final_feature_columns", "metric": col, "value": "selected"})
        except Exception as exc:
            rows.append({"section": "final_bundle", "metric": "joblib_load_error", "value": str(exc)})
    eval_dict = read_json(diagnostics_dir / "evaluation_dictionary_used.json")
    if eval_dict and not any(r["section"] == "cue_dictionary" for r in rows):
        for group, cues in (eval_dict.get("cue_groups") or {}).items():
            rows.append({"section": "cue_dictionary_from_diagnostics", "metric": group, "value": "; ".join(cues) if isinstance(cues, list) else norm(cues)})
    out["Classifier_Final_Details"] = pd.DataFrame(rows) if rows else pd.DataFrame({"note": ["No final classifier details found. Check classifier_dir path."]})

    candidates = []
    for base in [edir, cdir, plots_dir]:
        if base.exists():
            candidates += list(base.rglob("metrics_by_split.csv")) + list(base.rglob("threshold_metrics.csv"))
    for p in candidates:
        df = read_csv(p)
        if not df.empty:
            df.insert(0, "artifact", str(p))
            name = "Classifier_Model_Selection" if p.name == "metrics_by_split.csv" else "Classifier_Thresholds"
            out[name] = pd.concat([out.get(name, pd.DataFrame()), df], ignore_index=True)
    imp = read_csv(plots_dir / "classifier_feature_importance_top25.csv")
    if not imp.empty:
        out["Classifier_Feature_Importance"] = imp
    diffs = read_csv(plots_dir / "classifier_feature_mean_differences.csv")
    if not diffs.empty:
        out["Classifier_Feature_Diffs"] = diffs
    return out


def write_html_crosswalk(manual_xw: pd.DataFrame, llm_xw: pd.DataFrame, combined: pd.DataFrame, out: Path) -> None:
    def table(df: pd.DataFrame, n: int = 200) -> str:
        return "<p><i>Not available.</i></p>" if df.empty else df.head(n).to_html(index=False, escape=True)
    parts = ["<!doctype html><html><head><meta charset='utf-8'><style>body{font-family:Arial;margin:28px;color:#1f2937}h1{color:#0f172a}h2{color:#0f766e}.note{background:#f8fafc;border-left:4px solid #0ea5e9;padding:12px;margin:12px 0}table{border-collapse:collapse;font-size:12px;width:100%;margin-bottom:24px}td,th{border:1px solid #e5e7eb;padding:5px;vertical-align:top}th{background:#eff6ff}</style></head><body>"]
    parts.append("<h1>Source-model crosswalks to Manual V1 and LLM-induced V1</h1>")
    parts.append("<div class='note'>Scope: only DUO, ICO, ODRL, and FHIR source elements. LLM-induced mappings are evidence-derived from induction cards and fold-specific schemas, and require expert review.</div>")
    for title, df in [("Combined source element → Manual/LLM V1", combined), ("Source element → Manual V1", manual_xw), ("Source element → LLM-induced V1", llm_xw)]:
        parts.append(f"<h2>{html.escape(title)}</h2>{table(df)}")
    parts.append("</body></html>")
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text("".join(parts), encoding="utf-8")


def dataframe_for_sheet(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()
    # Drop columns that are entirely blank to reduce workbook clutter.
    work = df.copy()
    blank_cols = [c for c in work.columns if work[c].map(norm).eq("").all()]
    work = work.drop(columns=blank_cols)
    return work


def sheet_name(x: str) -> str:
    return re.sub(r"[\\/*?:\[\]]", "_", x)[:31]


def write_workbook(sheets: dict[str, pd.DataFrame], out_xlsx: Path) -> None:
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        for name, df in sheets.items():
            df = dataframe_for_sheet(df)
            if df.empty:
                continue
            df.to_excel(writer, sheet_name=sheet_name(name), index=False)
    wb = load_workbook(out_xlsx)
    header_fill = PatternFill("solid", fgColor="0F766E")
    header_font = Font(color="FFFFFF", bold=True)
    review_fill = PatternFill("solid", fgColor="FEF3C7")
    thin = Side(style="thin", color="D9E2EC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        review_cols = []
        for c in range(1, ws.max_column + 1):
            header = norm(ws.cell(1, c).value).lower()
            if header in {"expert_review_status", "expert_notes", "manual_expert_decision", "llm_expert_decision", "expert_meaning_preserved"}:
                review_cols.append(c)
            width = 14
            if any(k in header for k in ["definition", "details", "notes", "rationale", "examples", "criteria", "original", "reconstructed", "annotation", "cue", "value"]):
                width = 46
            elif any(k in header for k in ["field", "element", "source", "condition", "model"]):
                width = 28
            ws.column_dimensions[get_column_letter(c)].width = width
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border
                if cell.column in review_cols:
                    cell.fill = review_fill
        if review_cols and ws.max_row >= 2:
            dv = DataValidation(type="list", formula1='"keep,merge,split,rename,drop,context-dependent,needs discussion,yes,no,unclear"', allow_blank=True)
            ws.add_data_validation(dv)
            for c in review_cols:
                dv.add(f"{get_column_letter(c)}2:{get_column_letter(c)}{ws.max_row}")
    wb.save(out_xlsx)


def copytree_if_exists(src: Path, dst: Path) -> None:
    if src.exists():
        if dst.exists():
            shutil.rmtree(dst)
        shutil.copytree(src, dst)


def write_readme(package_dir: Path, sheets: dict[str, pd.DataFrame], overall: pd.DataFrame, classifier_method: pd.DataFrame) -> None:
    lines = [
        "# PI Expert Review Package v3",
        "",
        "This corrected package is organized around three review tasks:",
        "",
        "1. Review the meaning-preservation classifier development and finalized classifier.",
        "2. Compare modeling strategies using classifier score and holistic diagnostics.",
        "3. Review Manual V1 and LLM-induced V1 dictionaries and source-model crosswalks.",
        "",
        "## Start here",
        "",
        "- `expert_review_data_dictionary_and_crosswalks.xlsx` is the main expert-review workbook.",
        "- `expert_review_examples/expert_review_examples.html` contains fixed source-sentence examples with highlighted annotations.",
        "- `crosswalks/v1_crosswalk_review_summary.html` provides a browser-readable version of the source-model crosswalks.",
        "",
        "## Meaning-preservation classifier",
        "",
        classifier_method.to_markdown(index=False),
        "",
        "## Strategy-level metric snapshot",
        "",
        overall.to_markdown(index=False) if not overall.empty else "No metric summary found.",
        "",
        "## Workbook sheets generated",
        "",
    ]
    for k, df in sheets.items():
        if dataframe_for_sheet(df).empty:
            continue
        lines.append(f"- `{k}` ({len(df)} rows)")
    (package_dir / "README_PI_REVIEW_PACKAGE.md").write_text("\n".join(lines), encoding="utf-8")


def write_classifier_md(package_dir: Path, method: pd.DataFrame, details: pd.DataFrame, selection: pd.DataFrame) -> None:
    lines = ["# Meaning-Preservation Classifier Development", "", "## Method overview", "", method.to_markdown(index=False), "", "## Final classifier details", "", details.to_markdown(index=False) if not details.empty else "Not available.", "", "## Model-selection results", "", selection.head(50).to_markdown(index=False) if not selection.empty else "No `metrics_by_split.csv` found. Pass `--classifier_experiments_dir` if the classifier experiment outputs are stored elsewhere."]
    (package_dir / "CLASSIFIER_DEVELOPMENT_SUMMARY.md").write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--out_root", default="meta_model/functional_v1_experiments")
    ap.add_argument("--source_package_dir", default="", help="Existing v2 package to reuse examples/plots/diagnostics from. Defaults to <out_root>/pi_expert_review_package_v2.")
    ap.add_argument("--package_dir", default="", help="Output package. Defaults to <out_root>/pi_expert_review_package_v3.")
    ap.add_argument("--source_inventory_csv", default="meta_model/v0_union/source_element_inventory.csv")
    ap.add_argument("--manual_schema_yaml", default="meta_model/schemas/reduced_functional_v1_candidate.yaml")
    ap.add_argument("--manual_crosswalk_csv", default="meta_model/functional_v1/crosswalk/functional_v1_crosswalk.csv")
    ap.add_argument("--llm_induced_schema_root", default="meta_model/functional_v1/llm_induced")
    ap.add_argument("--evidence_cards_root", default="meta_model/functional_v1/llm_induction_cards")
    ap.add_argument("--llm_consensus_fields_csv", default="")
    ap.add_argument("--classifier_dir", default="meta_model/outputs/final_classifier")
    ap.add_argument("--classifier_experiments_dir", default="", help="Directory containing metrics_by_split.csv / threshold_metrics.csv from classifier development experiments.")
    ap.add_argument("--zip", action="store_true")
    args = ap.parse_args()

    out_root = Path(args.out_root)
    src_pkg = Path(args.source_package_dir) if args.source_package_dir else out_root / "pi_expert_review_package_v2"
    pkg = Path(args.package_dir) if args.package_dir else out_root / "pi_expert_review_package_v3"
    pkg.mkdir(parents=True, exist_ok=True)

    # Reuse high-quality existing folders from v2/current experiment root.
    for folder in ["expert_review_examples", "plots", "diagnostics", "comparison"]:
        copytree_if_exists(src_pkg / folder, pkg / folder)
        if not (pkg / folder).exists():
            copytree_if_exists(out_root / folder, pkg / folder)

    diagnostics_dir = pkg / "diagnostics"
    plots_dir = pkg / "plots"
    comparison_dir = pkg / "comparison"
    crosswalk_dir = pkg / "crosswalks"
    crosswalk_dir.mkdir(parents=True, exist_ok=True)

    raw_manual = read_csv(args.manual_crosswalk_csv)
    src0 = source_dictionary(args.source_inventory_csv, [raw_manual])
    look = make_lookup(src0)
    manual_xw = normalize_manual_crosswalk(args.manual_crosswalk_csv, look)
    src = source_dictionary(args.source_inventory_csv, [raw_manual, manual_xw])
    look = make_lookup(src)
    manual_xw = normalize_manual_crosswalk(args.manual_crosswalk_csv, look)
    llm_xw = build_llm_crosswalk(args.llm_induced_schema_root, args.evidence_cards_root, look, manual_xw)
    # Enrich lookup with LLM rows and refill manual/LLM labels if possible.
    src = source_dictionary(args.source_inventory_csv, [raw_manual, manual_xw, llm_xw])
    look = make_lookup(src)
    manual_xw = normalize_manual_crosswalk(args.manual_crosswalk_csv, look)
    llm_xw = build_llm_crosswalk(args.llm_induced_schema_root, args.evidence_cards_root, look, manual_xw)
    combined = combined_crosswalk(manual_xw, llm_xw)

    manual_xw.to_csv(crosswalk_dir / "manual_v1_source_model_crosswalk_for_review.csv", index=False)
    llm_xw.to_csv(crosswalk_dir / "llm_induced_v1_source_model_crosswalk_by_fold_for_review.csv", index=False)
    combined.to_csv(crosswalk_dir / "source_model_to_manual_and_llm_v1_crosswalk_for_review.csv", index=False)
    write_html_crosswalk(manual_xw, llm_xw, combined, crosswalk_dir / "v1_crosswalk_review_summary.html")

    method = classifier_method_sheet()
    classifier = classifier_artifacts(args.classifier_dir, args.classifier_experiments_dir, diagnostics_dir, plots_dir)
    overall = read_csv(comparison_dir / "schema_condition_overall.csv")
    by_llm = read_csv(comparison_dir / "schema_condition_by_llm.csv")
    by_model = read_csv(comparison_dir / "schema_condition_by_information_model.csv")
    cue = read_csv(diagnostics_dir / "cue_group_retention_summary_by_condition.csv")
    examples = read_csv(pkg / "expert_review_examples" / "expert_review_examples.csv")
    qualitative = read_csv(diagnostics_dir / "qualitative_relationship_error_review_sample.csv")

    sheets: dict[str, pd.DataFrame] = {
        "README": pd.DataFrame([
            {"item": "Purpose", "details": "One workbook for classifier review, modeling-strategy results, data dictionaries, crosswalks, and fixed examples."},
            {"item": "Recommended order", "details": "Classifier_Method -> Results_Overall -> Manual_V1_Dictionary / LLM_Induced_Dictionary -> Combined_Crosswalk -> Fixed_Examples."},
            {"item": "Crosswalk scope", "details": "Only source models DUO, ICO, ODRL, and FHIR are mapped to Manual V1 and LLM-induced V1."},
        ]),
        "Classifier_Method": method,
        "Classifier_Final_Details": classifier.get("Classifier_Final_Details", pd.DataFrame()),
        "Classifier_Model_Selection": classifier.get("Classifier_Model_Selection", pd.DataFrame()),
        "Classifier_Thresholds": classifier.get("Classifier_Thresholds", pd.DataFrame()),
        "Classifier_Feature_Importance": classifier.get("Classifier_Feature_Importance", pd.DataFrame()),
        "Classifier_Feature_Diffs": classifier.get("Classifier_Feature_Diffs", pd.DataFrame()),
        "Results_Overall": overall,
        "Results_by_LLM": by_llm,
        "Results_by_SourceModel": by_model,
        "Cue_Retention": cue,
        "Source_Dictionary_All": src,
        "Manual_V1_Dictionary": manual_dictionary(args.manual_schema_yaml),
        "LLM_Induced_Dictionary": llm_dictionary(args.llm_induced_schema_root, args.llm_consensus_fields_csv),
        "Crosswalk_Source_to_Manual": manual_xw,
        "Crosswalk_Source_to_LLM": llm_xw,
        "Combined_Crosswalk": combined,
        "Fixed_Examples": examples,
        "Qualitative_Errors": qualitative,
    }
    for sm in SOURCE_MODELS:
        sub = src[src["source_model"] == sm].copy() if not src.empty and "source_model" in src.columns else pd.DataFrame()
        if not sub.empty:
            sheets[f"Dict_{sm}"] = sub

    out_xlsx = pkg / "expert_review_data_dictionary_and_crosswalks.xlsx"
    write_workbook(sheets, out_xlsx)
    write_readme(pkg, sheets, overall, method)
    write_classifier_md(pkg, method, classifier.get("Classifier_Final_Details", pd.DataFrame()), classifier.get("Classifier_Model_Selection", pd.DataFrame()))

    if args.zip:
        z = shutil.make_archive(str(pkg), "zip", root_dir=pkg)
        print(f"Wrote zip: {z}")
    print(f"Wrote comprehensive PI package: {pkg}")
    print(f"Main workbook: {out_xlsx}")


if __name__ == "__main__":
    main()
