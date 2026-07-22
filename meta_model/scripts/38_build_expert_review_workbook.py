#!/usr/bin/env python
"""Build a single Excel workbook for expert review of consent meta-models.

The workbook is designed for PI/domain-expert adjudication. It combines:
- Manual V1 and LLM-induced V1 data dictionaries;
- source information-model dictionaries for DUO, ICO, ODRL, and FHIR;
- source-model -> Manual V1 / LLM-induced V1 crosswalks;
- performance summaries and classifier details;
- fixed annotation/reconstruction examples for review.

This script uses pandas plus openpyxl because it is intended to run inside the
project environment, not inside ChatGPT's artifact runtime.
"""
from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

import pandas as pd

try:
    import yaml
except Exception:  # pragma: no cover
    yaml = None  # type: ignore

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except Exception as exc:  # pragma: no cover
    raise SystemExit("openpyxl is required to build the expert-review workbook. Install with: pip install openpyxl") from exc

SOURCE_MODELS = ["DUO", "ICO", "ODRL", "FHIR"]
CONDITION_ORDER = [
    "individual_source_model_json",
    "union_v0_full_dictionary",
    "functional_v1_manual",
    "functional_v1_llm_induced",
    "functional_v1_llm_induced_consensus",
]


def norm(x: Any) -> str:
    if x is None:
        return ""
    try:
        if pd.isna(x):
            return ""
    except Exception:
        pass
    if isinstance(x, (list, tuple, set)):
        return "; ".join(norm(v) for v in x if norm(v))
    if isinstance(x, dict):
        return json.dumps(x, ensure_ascii=False)
    return " ".join(str(x).split())


def exists(path: str | Path | None) -> bool:
    return bool(path) and Path(path).exists()


def read_csv(path: str | Path | None) -> pd.DataFrame:
    if not exists(path):
        return pd.DataFrame()
    return pd.read_csv(path, low_memory=False).fillna("")


def read_json(path: str | Path | None) -> dict[str, Any]:
    if not exists(path):
        return {}
    try:
        return json.loads(Path(path).read_text())
    except Exception:
        return {}


def safe_sheet_name(name: str) -> str:
    name = re.sub(r"[\\/*?:\[\]]", "_", name)[:31]
    return name or "Sheet"


def source_model_canonical(x: Any) -> str:
    s = norm(x)
    u = s.upper()
    if "FHIR" in u:
        return "FHIR"
    if "ODRL" in u:
        return "ODRL"
    if "DUO" in u or s in {"GRU", "HMB", "NRES", "DS", "IRB", "TS"}:
        return "DUO"
    if "ICO" in u:
        return "ICO"
    return s


def tail(uid: Any) -> str:
    s = norm(uid)
    return s.split("::", 1)[1] if "::" in s else s


def pick(row: pd.Series | dict[str, Any], candidates: list[str]) -> str:
    for c in candidates:
        if isinstance(row, pd.Series):
            if c in row.index and norm(row.get(c)):
                return norm(row.get(c))
        else:
            if c in row and norm(row.get(c)):
                return norm(row.get(c))
    return ""


def listify(x: Any) -> str:
    if isinstance(x, str):
        s = x.strip()
        if not s:
            return ""
        try:
            obj = json.loads(s)
            if isinstance(obj, list):
                return "; ".join(norm(v) for v in obj if norm(v))
            if isinstance(obj, dict):
                return "; ".join(f"{k}: {norm(v)}" for k, v in obj.items())
        except Exception:
            return s
        return s
    if isinstance(x, list):
        return "; ".join(norm(v) for v in x if norm(v))
    if isinstance(x, dict):
        return "; ".join(f"{k}: {norm(v)}" for k, v in x.items())
    return norm(x)


def read_schema(path: str | Path | None) -> dict[str, Any]:
    if not exists(path):
        return {}
    p = Path(path)
    text = p.read_text()
    if p.suffix.lower() == ".json":
        return json.loads(text)
    if yaml is None:
        raise RuntimeError("PyYAML is required to read YAML schemas.")
    return yaml.safe_load(text)


def schema_fields(schema: dict[str, Any]) -> list[dict[str, Any]]:
    raw = schema.get("fields") or schema.get("schema_fields") or schema.get("functional_fields") or []
    rows: list[dict[str, Any]] = []
    if isinstance(raw, dict):
        for k, v in raw.items():
            d = dict(v) if isinstance(v, dict) else {"definition": v}
            d.setdefault("field_id", k)
            rows.append(d)
    elif isinstance(raw, list):
        for i, v in enumerate(raw):
            d = dict(v) if isinstance(v, dict) else {"field_id": str(v)}
            d.setdefault("field_id", d.get("id") or d.get("name") or f"field_{i:02d}")
            rows.append(d)
    return rows


def build_manual_dictionary(schema_yaml: str | Path | None) -> pd.DataFrame:
    schema = read_schema(schema_yaml)
    rows = []
    # Include sentence_decision separately if present in schema header.
    sd = schema.get("sentence_decision") or schema.get("decision") or {}
    if sd:
        rows.append({
            "strategy": "Manual Functional V1",
            "field_id": "sentence_decision",
            "field_name": "sentence_decision",
            "tier": "core",
            "definition": listify(sd.get("definition") or sd.get("description") or "Provision-level decision/polarity."),
            "allowed_values": listify(sd.get("allowed_values") or sd.get("values")),
            "include_when": "Provision-level sentence decision/polarity is expressed or inferable.",
            "exclude_when": "Do not use for local negation unless it changes the provision-level decision.",
            "examples": listify(sd.get("examples")),
            "expert_review_status": "",
            "expert_notes": "",
        })
    for f in schema_fields(schema):
        fid = pick(f, ["field_id", "id", "name"])
        rows.append({
            "strategy": "Manual Functional V1",
            "field_id": fid,
            "field_name": pick(f, ["name", "label", "field_id", "id"]) or fid,
            "tier": pick(f, ["tier", "status", "field_type"]),
            "definition": pick(f, ["definition", "description"]),
            "allowed_values": listify(f.get("allowed_values") or f.get("values")),
            "include_when": listify(f.get("include_when") or f.get("inclusion_criteria") or f.get("inclusion")),
            "exclude_when": listify(f.get("exclude_when") or f.get("exclusion_criteria") or f.get("exclusion")),
            "examples": listify(f.get("examples") or f.get("example_spans")),
            "expert_review_status": "",
            "expert_notes": "",
        })
    return pd.DataFrame(rows).drop_duplicates()


def field_evidence_ids(f: dict[str, Any]) -> str:
    ids: list[str] = []
    for k in ["evidence_card_ids", "assigned_evidence_cards", "supporting_evidence_cards", "evidence_cards", "source_cards", "cluster_ids", "stability_group_ids"]:
        v = f.get(k)
        if isinstance(v, list):
            ids.extend(norm(x) for x in v if norm(x))
        elif norm(v):
            ids.extend([x.strip() for x in re.split(r"[;|,]", norm(v)) if x.strip()])
    return "; ".join(dict.fromkeys(ids))


def build_llm_dictionary(schema_root: str | Path | None, consensus_fields_csv: str | Path | None = None) -> pd.DataFrame:
    rows = []
    root = Path(schema_root) if schema_root else Path("")
    if root.exists():
        for p in sorted(root.glob("fold_*/llm_induced_functional_v1_candidate.*")):
            if p.suffix.lower() not in {".yaml", ".yml", ".json"}:
                continue
            fold = p.parent.name
            schema = read_schema(p)
            for f in schema_fields(schema):
                fid = pick(f, ["field_id", "id", "name"])
                rows.append({
                    "schema_level": "fold_specific",
                    "fold": fold,
                    "field_id": fid,
                    "field_name": pick(f, ["name", "label", "field_id", "id"]) or fid,
                    "tier": pick(f, ["tier", "status", "field_type"]),
                    "definition": pick(f, ["definition", "description"]),
                    "allowed_values": listify(f.get("allowed_values") or f.get("values")),
                    "include_when": listify(f.get("include_when") or f.get("inclusion_criteria") or f.get("inclusion")),
                    "exclude_when": listify(f.get("exclude_when") or f.get("exclusion_criteria") or f.get("exclusion")),
                    "examples": listify(f.get("examples") or f.get("example_spans")),
                    "evidence_card_ids": field_evidence_ids(f),
                    "expert_review_status": "",
                    "expert_notes": "",
                })
    cons = read_csv(consensus_fields_csv)
    if not cons.empty:
        for _, r in cons.iterrows():
            rows.append({
                "schema_level": "post_cv_consensus",
                "fold": "consensus",
                "field_id": pick(r, ["consensus_field", "consensus_field_id", "field_id", "field", "name"]),
                "field_name": pick(r, ["consensus_field", "consensus_name", "name", "field"]),
                "tier": pick(r, ["tier", "status", "selection_tier"]),
                "definition": pick(r, ["definition", "description"]),
                "allowed_values": "",
                "include_when": pick(r, ["include_when", "inclusion_criteria"]),
                "exclude_when": pick(r, ["exclude_when", "exclusion_criteria"]),
                "examples": pick(r, ["examples", "top_spans", "example_spans"]),
                "evidence_card_ids": pick(r, ["evidence_card_ids", "supporting_fields", "fold_fields"]),
                "expert_review_status": "",
                "expert_notes": "",
            })
    return pd.DataFrame(rows).drop_duplicates()


def build_source_dictionary(inventory_csv: str | Path | None, manual_crosswalk: pd.DataFrame, llm_crosswalk: pd.DataFrame) -> pd.DataFrame:
    rows = []
    inv = read_csv(inventory_csv)
    if not inv.empty:
        for _, r in inv.iterrows():
            uid = pick(r, ["union_element_id", "source_element_id", "element_id", "id"])
            sm = source_model_canonical(pick(r, ["information_model", "source_model", "model", "canonical_information_model"]) or uid.split("::", 1)[0])
            if sm not in SOURCE_MODELS:
                continue
            rows.append({
                "source_model": sm,
                "source_element_id": uid,
                "source_element_tail": tail(uid),
                "source_element_label": pick(r, ["source_element_label", "label", "name", "source_label"]),
                "source_element_definition": pick(r, ["source_element_definition", "definition", "description"]),
                "source_element_uri_or_code": pick(r, ["uri", "url", "code", "concept_id"]),
                "source": "source_element_inventory",
            })
    for df, src in [(manual_crosswalk, "manual_crosswalk"), (llm_crosswalk, "llm_crosswalk")]:
        if df.empty:
            continue
        for _, r in df.iterrows():
            sm = source_model_canonical(pick(r, ["source_model", "information_model"])),
            sm = sm[0]
            if sm not in SOURCE_MODELS:
                continue
            sid = pick(r, ["source_element_id", "source_element", "union_element_id"])
            rows.append({
                "source_model": sm,
                "source_element_id": sid,
                "source_element_tail": tail(sid),
                "source_element_label": pick(r, ["source_element_label", "source_label", "label"]),
                "source_element_definition": pick(r, ["source_element_definition", "definition", "description"]),
                "source_element_uri_or_code": "",
                "source": src,
            })
    out = pd.DataFrame(rows).drop_duplicates()
    if not out.empty:
        out = out.sort_values(["source_model", "source_element_id", "source_element_label"])
    return out


def normalize_manual_crosswalk(path: str | Path | None, source_dict: pd.DataFrame | None = None) -> pd.DataFrame:
    df = read_csv(path)
    if df.empty:
        return df
    rows = []
    for _, r in df.iterrows():
        sm = source_model_canonical(pick(r, ["source_model", "information_model", "model"])),
        sm = sm[0]
        sid = pick(r, ["source_element_id", "source_element", "union_element_id", "element_id"])
        if sm not in SOURCE_MODELS:
            continue
        rows.append({
            "source_model": sm,
            "source_element_id": sid,
            "source_element_tail": tail(sid),
            "source_element_label": pick(r, ["source_element_label", "source_label", "label", "source_element"]),
            "source_element_definition": pick(r, ["source_element_definition", "definition", "description"]),
            "manual_v1_field": pick(r, ["manual_v1_field", "v1_field", "functional_field", "proposed_v1_field", "target_field"]),
            "secondary_manual_v1_fields": pick(r, ["secondary_v1_fields_json", "secondary_fields"]),
            "mapping_type": pick(r, ["mapping_type", "manual_mapping_type", "relationship", "mapping_relation"]),
            "context_rule_or_rationale": pick(r, ["context_rule", "rationale", "rule", "notes"]),
            "requires_context_review": pick(r, ["requires_context_review"]),
            "expert_review_status": "",
            "expert_notes": "",
        })
    out = pd.DataFrame(rows).drop_duplicates()
    return fill_source_details(out, source_dict)


def normalize_llm_crosswalk(path: str | Path | None, source_dict: pd.DataFrame | None = None) -> pd.DataFrame:
    df = read_csv(path)
    if df.empty:
        return df
    rows = []
    for _, r in df.iterrows():
        sm = source_model_canonical(pick(r, ["source_model", "information_model", "model"])),
        sm = sm[0]
        sid = pick(r, ["source_element_id", "source_element", "union_element_id", "element_id"])
        if sm not in SOURCE_MODELS:
            continue
        rows.append({
            "fold": pick(r, ["fold", "fold_id"]),
            "source_model": sm,
            "source_element_id": sid,
            "source_element_tail": tail(sid),
            "source_element_label": pick(r, ["source_element_label", "source_label", "label", "source_element"]),
            "source_element_definition": pick(r, ["source_element_definition", "definition", "description"]),
            "llm_induced_v1_field": pick(r, ["llm_induced_v1_field", "llm_induced_field", "field_id"]),
            "llm_induced_v1_field_name": pick(r, ["llm_induced_v1_field_name", "llm_induced_field_name", "field_name"]),
            "llm_induced_v1_definition": pick(r, ["llm_induced_v1_definition", "llm_induced_definition", "definition"]),
            "manual_v1_fields_linked_by_source_element": pick(r, ["manual_v1_fields_linked_by_source_element", "manual_v1_field"]),
            "mapping_basis": pick(r, ["llm_mapping_basis", "mapping_basis"]),
            "evidence_card_ids": pick(r, ["evidence_card_ids"]),
            "expert_review_status": "",
            "expert_notes": "",
        })
    out = pd.DataFrame(rows).drop_duplicates()
    return fill_source_details(out, source_dict)


def fill_source_details(df: pd.DataFrame, source_dict: pd.DataFrame | None) -> pd.DataFrame:
    if df.empty or source_dict is None or source_dict.empty:
        return df
    lookup = {}
    for _, r in source_dict.iterrows():
        for key in [norm(r.get("source_element_id")).lower(), tail(r.get("source_element_id")).lower(), norm(r.get("source_element_label")).lower()]:
            if key:
                lookup.setdefault(key, r.to_dict())
    rows = []
    for _, r in df.iterrows():
        d = r.to_dict()
        key = norm(d.get("source_element_id")).lower() or tail(d.get("source_element_id")).lower()
        match = lookup.get(key) or lookup.get(tail(d.get("source_element_id")).lower()) or {}
        for c in ["source_element_label", "source_element_definition"]:
            if not norm(d.get(c)) and match:
                d[c] = norm(match.get(c))
        rows.append(d)
    return pd.DataFrame(rows)


def combined_crosswalk(manual: pd.DataFrame, llm: pd.DataFrame) -> pd.DataFrame:
    keys = ["source_model", "source_element_id", "source_element_tail", "source_element_label", "source_element_definition"]
    if manual.empty and llm.empty:
        return pd.DataFrame()
    parts = []
    if not manual.empty:
        m = manual.groupby(keys, dropna=False).agg(
            manual_v1_fields=("manual_v1_field", lambda x: "; ".join(sorted(set(norm(v) for v in x if norm(v))))),
            manual_mapping_types=("mapping_type", lambda x: "; ".join(sorted(set(norm(v) for v in x if norm(v))))),
        ).reset_index()
        parts.append(m)
    if not llm.empty:
        l = llm.groupby(keys, dropna=False).agg(
            llm_induced_v1_fields=("llm_induced_v1_field", lambda x: "; ".join(sorted(set(norm(v) for v in x if norm(v))))),
            llm_induced_v1_field_names=("llm_induced_v1_field_name", lambda x: "; ".join(sorted(set(norm(v) for v in x if norm(v))))),
            llm_induced_folds=("fold", lambda x: "; ".join(sorted(set(norm(v) for v in x if norm(v))))),
            llm_mapping_basis=("mapping_basis", lambda x: "; ".join(sorted(set(norm(v) for v in x if norm(v))))),
        ).reset_index()
        parts.append(l)
    if len(parts) == 1:
        out = parts[0]
    else:
        out = parts[0].merge(parts[1], on=keys, how="outer")
    out["manual_expert_decision"] = ""
    out["llm_expert_decision"] = ""
    out["expert_notes"] = ""
    return out.sort_values(["source_model", "source_element_id"])


def classifier_details(summary_json: str | Path | None, diagnostics_dir: Path) -> pd.DataFrame:
    rows = []
    summary = read_json(summary_json)
    if summary:
        def flatten(prefix: str, obj: Any) -> None:
            if isinstance(obj, dict):
                for k, v in obj.items():
                    flatten(f"{prefix}.{k}" if prefix else str(k), v)
            elif isinstance(obj, list):
                rows.append({"section": "training_summary", "metric": prefix, "value": "; ".join(norm(v) for v in obj)})
            else:
                rows.append({"section": "training_summary", "metric": prefix, "value": norm(obj)})
        flatten("", summary)
    eval_dict = read_json(diagnostics_dir / "evaluation_dictionary_used.json")
    if eval_dict:
        cue_groups = eval_dict.get("cue_groups", {})
        for group, cues in cue_groups.items():
            rows.append({"section": "cue_dictionary", "metric": group, "value": "; ".join(cues) if isinstance(cues, list) else norm(cues)})
    return pd.DataFrame(rows)


def readme_sheet(package_dir: Path) -> pd.DataFrame:
    return pd.DataFrame([
        {"item": "Purpose", "details": "One workbook for PI/domain-expert review of source dictionaries, Manual V1, LLM-induced V1, crosswalks, metrics, and examples."},
        {"item": "Recommended first pass", "details": "Review Manual_V1_Dictionary and LLM_Induced_Dictionary, then Combined_Crosswalk, then Fixed_Examples."},
        {"item": "Crosswalk scope", "details": "Only source information models DUO, ICO, ODRL, and FHIR are mapped to Manual V1 and LLM-induced V1."},
        {"item": "Expert decision columns", "details": "Use expert_review_status/manual_expert_decision/llm_expert_decision/expert_notes to mark keep, merge, split, rename, drop, context-dependent, or needs discussion."},
        {"item": "Package directory", "details": str(package_dir)},
    ])


def write_workbook(sheets: dict[str, pd.DataFrame], out_xlsx: Path) -> None:
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        for name, df in sheets.items():
            if df is None or df.empty:
                df = pd.DataFrame({"note": ["No data available for this sheet."]})
            df.to_excel(writer, sheet_name=safe_sheet_name(name), index=False)

    wb = load_workbook(out_xlsx)
    header_fill = PatternFill("solid", fgColor="0F766E")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2EC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    review_fill = PatternFill("solid", fgColor="FEF3C7")
    readme_fill = PatternFill("solid", fgColor="EFF6FF")

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border
                if str(ws.cell(1, cell.column).value).lower() in {"expert_review_status", "expert_notes", "manual_expert_decision", "llm_expert_decision", "expert_meaning_preserved"}:
                    cell.fill = review_fill
        if ws.title == "README":
            for row in ws.iter_rows():
                for cell in row:
                    cell.fill = readme_fill if cell.row > 1 else header_fill
        max_cols = min(ws.max_column, 40)
        for col_idx in range(1, max_cols + 1):
            col_letter = get_column_letter(col_idx)
            header = norm(ws.cell(1, col_idx).value).lower()
            width = 14
            if any(k in header for k in ["definition", "details", "notes", "examples", "criteria", "rationale", "original", "reconstructed", "annotation", "dictionary"]):
                width = 42
            elif any(k in header for k in ["field", "element", "source", "condition"]):
                width = 26
            elif any(k in header for k in ["score", "rate", "count", "fold", "tier"]):
                width = 14
            ws.column_dimensions[col_letter].width = width
        ws.row_dimensions[1].height = 28

        # Add dropdown review status to expert review columns.
        review_cols = []
        for c in range(1, ws.max_column + 1):
            if norm(ws.cell(1, c).value).lower() in {"expert_review_status", "manual_expert_decision", "llm_expert_decision", "expert_meaning_preserved"}:
                review_cols.append(c)
        if review_cols and ws.max_row >= 2:
            dv = DataValidation(type="list", formula1='"keep,merge,split,rename,drop,context-dependent,needs discussion,yes,no,unclear"', allow_blank=True)
            ws.add_data_validation(dv)
            for c in review_cols:
                dv.add(f"{get_column_letter(c)}2:{get_column_letter(c)}{ws.max_row}")
    wb.save(out_xlsx)


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--package_dir", required=True, help="PI package directory, usually <OUT_ROOT>/pi_expert_review_package_v2")
    ap.add_argument("--out_xlsx", default="", help="Default: <package_dir>/expert_review_data_dictionary_and_crosswalks.xlsx")
    ap.add_argument("--source_inventory_csv", default="meta_model/v0_union/source_element_inventory.csv")
    ap.add_argument("--manual_schema_yaml", default="meta_model/schemas/reduced_functional_v1_candidate.yaml")
    ap.add_argument("--llm_induced_schema_root", default="meta_model/functional_v1/llm_induced")
    ap.add_argument("--llm_consensus_fields_csv", default="")
    ap.add_argument("--classifier_summary_json", default="meta_model/outputs/final_classifier/final_classifier_training_summary.json")
    args = ap.parse_args()

    package_dir = Path(args.package_dir)
    out_xlsx = Path(args.out_xlsx) if args.out_xlsx else package_dir / "expert_review_data_dictionary_and_crosswalks.xlsx"
    crosswalk_dir = package_dir / "crosswalks"
    diagnostics_dir = package_dir / "diagnostics"
    comparison_dir = package_dir / "comparison"
    examples_dir = package_dir / "expert_review_examples"

    # First load crosswalks generated by v2 package builder, then enrich from source inventory.
    raw_manual_xw = read_csv(crosswalk_dir / "manual_v1_source_model_crosswalk_for_review.csv")
    raw_llm_xw = read_csv(crosswalk_dir / "llm_induced_v1_source_model_crosswalk_by_fold_for_review.csv")
    source_dict_pre = build_source_dictionary(args.source_inventory_csv, raw_manual_xw, raw_llm_xw)
    manual_xw = normalize_manual_crosswalk(crosswalk_dir / "manual_v1_source_model_crosswalk_for_review.csv", source_dict_pre)
    llm_xw = normalize_llm_crosswalk(crosswalk_dir / "llm_induced_v1_source_model_crosswalk_by_fold_for_review.csv", source_dict_pre)
    source_dict = build_source_dictionary(args.source_inventory_csv, manual_xw, llm_xw)
    manual_xw = fill_source_details(manual_xw, source_dict)
    llm_xw = fill_source_details(llm_xw, source_dict)
    combined = combined_crosswalk(manual_xw, llm_xw)

    metric_summary = read_csv(comparison_dir / "schema_condition_overall.csv")
    strategy_llm = read_csv(comparison_dir / "schema_condition_by_llm.csv")
    strategy_model = read_csv(comparison_dir / "schema_condition_by_information_model.csv")
    fixed_examples = read_csv(examples_dir / "expert_review_examples.csv")
    qualitative = read_csv(diagnostics_dir / "qualitative_relationship_error_review_sample.csv")
    classifier = classifier_details(args.classifier_summary_json, diagnostics_dir)
    manual_dict = build_manual_dictionary(args.manual_schema_yaml)
    llm_dict = build_llm_dictionary(args.llm_induced_schema_root, args.llm_consensus_fields_csv)

    sheets: dict[str, pd.DataFrame] = {
        "README": readme_sheet(package_dir),
        "Metric_Summary": metric_summary,
        "Strategy_x_LLM": strategy_llm,
        "Strategy_x_SourceModel": strategy_model,
        "Classifier_Details": classifier,
        "Source_Dictionaries": source_dict,
        "Manual_V1_Dictionary": manual_dict,
        "LLM_Induced_Dictionary": llm_dict,
        "Crosswalk_Source_to_Manual": manual_xw,
        "Crosswalk_Source_to_LLM": llm_xw,
        "Combined_Crosswalk": combined,
        "Fixed_Examples": fixed_examples,
        "Qualitative_Errors": qualitative,
    }
    for sm in SOURCE_MODELS:
        sm_df = source_dict[source_dict["source_model"] == sm].copy() if not source_dict.empty and "source_model" in source_dict.columns else pd.DataFrame()
        sheets[f"Dict_{sm}"] = sm_df

    write_workbook(sheets, out_xlsx)
    print(f"Wrote expert-review workbook: {out_xlsx}")


if __name__ == "__main__":
    main()
